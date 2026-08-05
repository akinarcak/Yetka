# -*- coding: utf-8 -*-
"""Tests for the workbook serializer that replaces openpyxl's removed one.

`save_virtual_workbook` was deprecated in openpyxl 3.0 and removed in 3.1, and
three call sites depended on it. The replacement is small, so what is worth
asserting is not that it runs but that it produces the same workbook: an export
that silently loses a sheet or a row would look like a working download.

The differential test runs only while a version still carrying
`save_virtual_workbook` is installed, and skips itself afterwards.
"""
import unittest
from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook

from common.drf.renders.excel import workbook_to_bytes

try:  # openpyxl < 3.1
    from openpyxl.writer.excel import save_virtual_workbook
    HAS_LEGACY_SAVER = True
except ImportError:  # pragma: no cover - taken once the dependency moves
    HAS_LEGACY_SAVER = False

ROWS = [
    ['name', 'address', 'comment'],
    ['web-01', '10.0.0.1', 'ascii only'],
    ['sunucu-02', '10.0.0.2', 'Türkçe karakterler: ışğüçö'],
    ['', '', ''],
    ['quote"', 'back\\slash', 'tab\tand newline'],
]


def _build():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Assets'
    for row in ROWS:
        sheet.append(row)
    second = workbook.create_sheet('Second')
    second.append(['only', 'one', 'row'])
    return workbook


def _read_back(payload):
    loaded = load_workbook(BytesIO(payload))
    return {
        name: [list(row) for row in loaded[name].iter_rows(values_only=True)]
        for name in loaded.sheetnames
    }


class WorkbookToBytesTests(SimpleTestCase):
    def test_returns_a_real_xlsx(self):
        payload = workbook_to_bytes(_build())
        self.assertIsInstance(payload, bytes)
        # xlsx is a zip; PK is its magic number.
        self.assertTrue(payload.startswith(b'PK'))

    def test_every_sheet_and_cell_survives(self):
        payload = workbook_to_bytes(_build())
        read_back = _read_back(payload)
        self.assertEqual(sorted(read_back), ['Assets', 'Second'])
        self.assertEqual(read_back['Assets'][0], ['name', 'address', 'comment'])
        self.assertEqual(read_back['Assets'][2][2], 'Türkçe karakterler: ışğüçö')
        self.assertEqual(len(read_back['Assets']), len(ROWS))

    def test_serializing_twice_does_not_change_the_workbook(self):
        # The renderer calls this once, but a workbook that is consumed by
        # being written would be a trap for any future caller.
        workbook = _build()
        first = _read_back(workbook_to_bytes(workbook))
        second = _read_back(workbook_to_bytes(workbook))
        self.assertEqual(first, second)


@unittest.skipUnless(
    HAS_LEGACY_SAVER,
    'openpyxl no longer ships save_virtual_workbook; the comparison is moot',
)
class DifferentialAgainstOpenpyxlTests(SimpleTestCase):
    def test_same_content_as_save_virtual_workbook(self):
        # The bytes cannot be compared directly -- xlsx is a zip and carries
        # timestamps -- so the workbooks are compared after reading them back.
        self.assertEqual(
            _read_back(workbook_to_bytes(_build())),
            _read_back(save_virtual_workbook(_build())),
        )

    def test_same_archive_members(self):
        import zipfile

        def members(payload):
            with zipfile.ZipFile(BytesIO(payload)) as archive:
                return sorted(archive.namelist())

        self.assertEqual(
            members(workbook_to_bytes(_build())),
            members(save_virtual_workbook(_build())),
        )
